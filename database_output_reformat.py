import os
import openpyxl

print(os.getcwd())
HOME_DIR = os.path.join(os.getcwd(), 'Desktop', 'leopards_output')
LOCATION_SHORTHANDS = {
    'Washington, District of Columbia': 'DC',
    'New York, New York': 'NY',
    'Washington, DC': 'DC',
    'Chicago, Illinois': 'CHI',
    'Silicon Valley': 'SV',
    'Palo Alto, California': 'PA',
    'San Francisco, California': 'SF',
    'Los Angeles, California': 'LA',
    'Seattle, Washington': 'WA',
    'San Diego, California': 'SD',
    'London, England': 'UK',
    'Boston, Massachusetts': 'BOS',
    'Philadelphia, Pennsylvania': 'PHL'
    }

def get_workbooks() -> list[str]:
    workbooks = []
    for file in os.listdir(HOME_DIR):
        if file.endswith('.xlsx') and not file.startswith('FORMATTED_'):
            workbooks.append(os.path.join(os.getcwd(), HOME_DIR, file))
    return workbooks


def delete_columns(ws, columns_to_keep):
    header = [cell.value for cell in ws[1]]
    for i in range(len(header)-1, -1, -1):
        col_name = header[i]
        if col_name not in columns_to_keep:
            ws.delete_cols(header.index(col_name)+1)

def change_column_names(ws, names_mapping):
    for cell in ws[1]:
        cell.value = names_mapping.get(cell.value, cell.value)

def format_columns(ws):
    header = {cell.value : cell.column for cell in ws[1]}
    firm_col = header['Firm']
    loc_col = header['Loc']
    bio_col = header['Bio']
    linkedin_col = header['LinkedIn']
    origin_exit_col = header['Origin Firm/ Exit Firm']
    for row in ws.iter_rows(min_row=2):
        firm_cell = row[firm_col-1]
        firm_cell.value = firm_cell.value.split()[0].strip(',').strip()
        origin_exit_cell = row[origin_exit_col-1]
        if origin_exit_cell.value:
            origin_exit_cell.value = origin_exit_cell.value.split()[0].strip(',').strip()
        bio_cell = row[bio_col-1]
        if bio_cell.hyperlink:
            bio_cell.value = 'Bio'
        linkedin_cell = row[linkedin_col-1]
        if linkedin_cell.hyperlink:
            linkedin_cell.value = 'LinkedIn'
        loc_cell = row[loc_col-1]
        for key, val in LOCATION_SHORTHANDS.items():
            loc_cell.value = loc_cell.value.replace(key, val)


def set_print_setup(ws):
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    
    ws.print_options.gridLines = True
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = True
    ws.print_options.headings = True
    
    ws.oddHeader.left.text = '&F'



def set_width(ws):
    TARGET_WIDTH = 130
    total_width = 0
    for col in ws.columns:
        # print(col[0].value)
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
                max_length = min(60, max_length)
        
        width = max_length + 2
        ws.column_dimensions[col_letter].width = width
        total_width += width

        # if col[0].value == 'Notes - Prior Firm':
        #     print(total_width)
        #     ws.column_dimensions[col_letter].width += (TARGET_WIDTH - total_width)



def open_workbook(workbook_file: str):
    columns_to_keep = {'Firm', 'Last', 'First', 'JD / JD Equivalent', 'Law School', 'Loc', 'Link', 'LinkedIn', 'Notes', 'Origin Firm/ Exit Firm'}
    column_names = {'JD / JD Equivalent': 'JD',
                    'Link': 'Bio',
                    'Notes': 'Notes - Prior Firm'}
    wb = openpyxl.load_workbook(workbook_file)
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        delete_columns(ws, columns_to_keep)
        change_column_names(ws, column_names)
        format_columns(ws)

        set_width(ws)

        set_print_setup(ws)

    file_path, file = os.path.split(workbook_file)
    wb.save(os.path.join(file_path, f'FORMATTED_{file}'))


def main():
    workbooks = get_workbooks()
    for wb in workbooks:
        open_workbook(wb)

if __name__ == '__main__':
    main()
